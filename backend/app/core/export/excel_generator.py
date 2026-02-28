"""
Excel report generator for lifetime prediction data.

Generates Excel files with:
- Multiple sheets (Summary, Parameters, Results, Cycles)
- Formatted headers and data
- Embedded charts
- Data validation
"""
import logging
from io import BytesIO
from datetime import datetime
from typing import Dict, Any, Optional, List, Tuple
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side
)
from openpyxl.chart import (
    BarChart, LineChart, Reference
)
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np


logger = logging.getLogger(__name__)


# Model equation references (text format for Excel)
MODEL_EQUATIONS = {
    "coffin-manson": "Nf = A * (delta_Tj)^(-alpha)",
    "coffin-manson-arrhenius": "Nf = A * (delta_Tj)^(-alpha) * exp(Ea / (k * Tj))",
    "norris-landzberg": "Nf = A * (delta_Tj)^(-alpha) * f^beta * exp(Ea / (k * Tj_max))",
    "cips-2008": "Nf = K * (delta_Tj)^beta1 * exp(beta2 / (Tj_m + 273)) * t_on^beta3 * I^beta4 * V^beta5 * D^beta6",
    "lesit": "Nf = K * (delta_Tj)^(-alpha) * exp(Ea / (k * Tj_min))",
}

MODEL_DESCRIPTIONS = {
    "coffin-manson": "Basic Coffin-Manson model considering only temperature swing amplitude.",
    "coffin-manson-arrhenius": "Coffin-Manson model with Arrhenius term for mean temperature effects.",
    "norris-landzberg": "Norris-Landzberg model including frequency and maximum temperature effects.",
    "cips-2008": "CIPS 2008 (Bayerer) model with comprehensive stress factors including current, voltage, and bond wire diameter.",
    "lesit": "LESIT model focusing on temperature swing and minimum temperature effects.",
}


@dataclass
class ExcelConfig:
    """Configuration for Excel report generation."""
    author: str = "IOL PC Research System"
    title: str = "Lifetime Prediction Report"
    include_charts: bool = True
    include_formulas: bool = False  # Add Excel formulas for calculations
    freeze_headers: bool = True  # Freeze header rows


class ExcelGenerator:
    """
    Excel report generator for lifetime prediction results.

    Generates multi-sheet Excel workbooks with formatted data,
    charts, and optional formulas.
    """

    # Style definitions
    HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    TITLE_FONT = Font(bold=True, size=14, color="000000")
    SUBTITLE_FONT = Font(bold=True, size=12, color="4472C4")

    NORMAL_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    THICK_BORDER = Border(
        left=Side(style='medium', color='000000'),
        right=Side(style='medium', color='000000'),
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    def __init__(self, config: Optional[ExcelConfig] = None):
        """
        Initialize Excel generator.

        Args:
            config: Excel configuration options
        """
        self.config = config or ExcelConfig()

    def generate_lifetime_report(
        self,
        prediction: Dict[str, Any],
        parameters: Dict[str, Any],
        mission_profile: Optional[Dict[str, Any]] = None,
        rainflow_cycles: Optional[List[Dict[str, Any]]] = None,
        output_path: Optional[str] = None
    ) -> bytes:
        """
        Generate a complete lifetime prediction Excel report.

        Args:
            prediction: Prediction result data
            parameters: Model input parameters
            mission_profile: Mission profile data (optional)
            rainflow_cycles: Rainflow cycle counting results (optional)
            output_path: Optional file path to save the Excel file

        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create all sheets
        self._create_summary_sheet(wb, prediction, parameters)
        self._create_parameters_sheet(wb, parameters)
        self._create_results_sheet(wb, prediction, parameters)

        if rainflow_cycles:
            self._create_cycles_sheet(wb, rainflow_cycles)

        if mission_profile:
            self._create_mission_profile_sheet(wb, mission_profile)

        if self.config.include_charts:
            self._create_charts_sheet(wb, prediction, parameters)

        # Set workbook properties
        wb.properties.author = self.config.author
        wb.properties.title = self.config.title
        wb.properties.created = datetime.now()

        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Save to file if path provided
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())

        return buffer.getvalue()

    def _create_summary_sheet(
        self,
        wb: Workbook,
        prediction: Dict[str, Any],
        parameters: Dict[str, Any]
    ):
        """Create the summary sheet with key results."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['B1'] = self.config.title
        ws['B1'].font = self.TITLE_FONT
        ws.merge_cells('B1:E1')

        # Subtitle
        prediction_name = prediction.get('name', 'Unnamed Prediction')
        ws['B2'] = f"Analysis: {prediction_name}"
        ws['B2'].font = self.SUBTITLE_FONT
        ws.merge_cells('B2:E2')

        # Timestamp
        ws['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['B3'].font = Font(size=9, italic=True, color="666666")
        ws.merge_cells('B2:E3')

        row = 5

        # Model Information section
        ws[f'B{row}'] = "Model Information"
        ws[f'B{row}'].font = self.SUBTITLE_FONT
        row += 2

        model_type = prediction.get('model_type', 'Unknown')
        ws[f'B{row}'] = "Model Type:"
        ws[f'C{row}'] = self._format_model_name(model_type)
        row += 1

        ws[f'B{row}'] = "Description:"
        ws[f'C{row}'] = MODEL_DESCRIPTIONS.get(model_type, "Custom model")
        row += 2

        # Key Results section
        ws[f'B{row}'] = "Key Results"
        ws[f'B{row}'].font = self.SUBTITLE_FONT
        row += 1

        # Results table
        headers = ["Result", "Value", "Unit", "Status"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.NORMAL_BORDER

        row += 1

        # Add results
        lifetime_years = prediction.get('predicted_lifetime_years')
        lifetime_cycles = prediction.get('predicted_lifetime_cycles')
        damage = prediction.get('total_damage')
        confidence = prediction.get('confidence_level')

        results_data = []

        if lifetime_years is not None:
            status = self._get_lifetime_status_text(lifetime_years, 'years')
            results_data.append([
                "Predicted Lifetime",
                f"{lifetime_years:.2f}",
                "years",
                status
            ])

        if lifetime_cycles is not None:
            status = self._get_lifetime_status_text(lifetime_cycles, 'cycles')
            results_data.append([
                "Cycles to Failure",
                f"{lifetime_cycles:,.0f}",
                "cycles",
                status
            ])

        if damage is not None:
            status = self._get_damage_status_text(damage)
            results_data.append([
                "Total Damage",
                f"{damage:.4f}",
                "-",
                status
            ])

        if confidence is not None:
            ci_lower = prediction.get('confidence_interval_lower')
            ci_upper = prediction.get('confidence_interval_upper')
            if ci_lower is not None and ci_upper is not None:
                results_data.append([
                    f"{confidence:.0%} Confidence Interval",
                    f"[{ci_lower:.2f}, {ci_upper:.2f}]",
                    "years",
                    "OK"
                ])

        # Write results
        for result_row in results_data:
            for col, value in enumerate(result_row, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.NORMAL_BORDER

                # Color code status
                if col == 5:  # Status column
                    if value == "CRITICAL":
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif value == "WARNING":
                        cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                    elif value == "OK":
                        cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            row += 1

        # Adjust column widths
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15

        # Freeze header if configured
        if self.config.freeze_headers:
            ws.freeze_panes = 'B7'

    def _create_parameters_sheet(self, wb: Workbook, parameters: Dict[str, Any]):
        """Create the input parameters sheet."""
        ws = wb.create_sheet("Parameters")

        # Title
        ws['B1'] = "Input Parameters"
        ws['B1'].font = self.TITLE_FONT
        ws.merge_cells('B1:D1')

        row = 3

        # Define parameter categories
        categories = {
            "Thermal Parameters": {
                'delta_Tj': ('Junction Temperature Swing', '°C'),
                'Tj_max': ('Maximum Junction Temperature', '°C'),
                'Tj_mean': ('Mean Junction Temperature', '°C'),
                'Tj_min': ('Minimum Junction Temperature', '°C'),
                'Tc': ('Case Temperature', '°C'),
                'Ta': ('Ambient Temperature', '°C'),
                'Rth': ('Thermal Resistance', 'K/W'),
                'Cth': ('Thermal Capacitance', 'J/K'),
            },
            "Electrical Parameters": {
                'I': ('Current', 'A'),
                'V': ('Voltage', 'V'),
                'I_RMS': ('RMS Current', 'A'),
                'f': ('Switching Frequency', 'Hz'),
                'power': ('Power', 'W'),
            },
            "Timing Parameters": {
                't_on': ('On Time', 's'),
                't_off': ('Off Time', 's'),
                'period': ('Cycle Period', 's'),
                'duty_cycle': ('Duty Cycle', '%'),
            },
            "Physical Parameters": {
                'D': ('Bond Wire Diameter', 'µm'),
                'chip_area': ('Chip Area', 'mm²'),
                'solder_area': ('Solder Area', 'mm²'),
            },
            "Environmental Parameters": {
                'humidity': ('Humidity', '%RH'),
                'vibration': ('Vibration', 'g'),
                'altitude': ('Altitude', 'm'),
            },
        }

        # Create tables for each category
        for category_name, category_params in categories.items():
            # Category header
            ws[f'B{row}'] = category_name
            ws[f'B{row}'].font = self.SUBTITLE_FONT
            row += 1

            # Table header
            headers = ["Parameter", "Value", "Unit", "Description"]
            for col, header in enumerate(headers, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.NORMAL_BORDER

            row += 1
            found_params = False

            for key, (name, unit) in category_params.items():
                if key in parameters:
                    found_params = True
                    value = parameters[key]
                    formatted_value = self._format_value(value)

                    ws.cell(row=row, column=2, value=name)
                    ws.cell(row=row, column=3, value=formatted_value)
                    ws.cell(row=row, column=4, value=unit)

                    # Add border
                    for col in range(2, 6):
                        ws.cell(row=row, column=col).border = self.NORMAL_BORDER

                    row += 1

            if found_params:
                row += 1

        # Add any uncategorized parameters
        categorized_keys = set()
        for category_params in categories.values():
            categorized_keys.update(category_params.keys())

        uncategorized = {k: v for k, v in parameters.items() if k not in categorized_keys}

        if uncategorized:
            ws[f'B{row}'] = "Other Parameters"
            ws[f'B{row}'].font = self.SUBTITLE_FONT
            row += 1

            headers = ["Parameter", "Value"]
            for col, header in enumerate(headers, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.NORMAL_BORDER

            row += 1

            for key, value in uncategorized.items():
                formatted_value = self._format_value(value)
                ws.cell(row=row, column=2, value=key)
                ws.cell(row=row, column=3, value=formatted_value)

                for col in range(2, 4):
                    ws.cell(row=row, column=col).border = self.NORMAL_BORDER

                row += 1

        # Adjust column widths
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30

        # Freeze header
        if self.config.freeze_headers:
            ws.freeze_panes = 'B4'

    def _create_results_sheet(
        self,
        wb: Workbook,
        prediction: Dict[str, Any],
        parameters: Dict[str, Any]
    ):
        """Create the detailed results sheet."""
        ws = wb.create_sheet("Results")

        # Title
        ws['B1'] = "Prediction Results"
        ws['B1'].font = self.TITLE_FONT
        ws.merge_cells('B1:E1')

        row = 3

        # Model Equation section
        ws[f'B{row}'] = "Model Equation"
        ws[f'B{row}'].font = self.SUBTITLE_FONT
        row += 1

        model_type = prediction.get('model_type', 'Unknown')
        equation = MODEL_EQUATIONS.get(model_type, "Custom equation")
        ws[f'B{row}'] = equation
        ws[f'B{row}'].font = Font(italic=True, size=10)
        row += 2

        # Detailed results table
        headers = ["Parameter", "Value", "Lower Bound", "Upper Bound", "Unit"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.NORMAL_BORDER

        row += 1

        # Results data
        results_data = []

        lifetime_years = prediction.get('predicted_lifetime_years')
        if lifetime_years is not None:
            ci_lower = prediction.get('confidence_interval_lower')
            ci_upper = prediction.get('confidence_interval_upper')
            results_data.append([
                "Predicted Lifetime",
                lifetime_years,
                ci_lower if ci_lower else "N/A",
                ci_upper if ci_upper else "N/A",
                "years"
            ])

        lifetime_cycles = prediction.get('predicted_lifetime_cycles')
        if lifetime_cycles is not None:
            results_data.append([
                "Cycles to Failure",
                lifetime_cycles,
                "N/A",
                "N/A",
                "cycles"
            ])

        damage = prediction.get('total_damage')
        if damage is not None:
            results_data.append([
                "Total Damage",
                damage,
                "N/A",
                "N/A",
                "-"
            ])

        confidence = prediction.get('confidence_level')
        if confidence is not None:
            results_data.append([
                "Confidence Level",
                f"{confidence:.1%}",
                "N/A",
                "N/A",
                "-"
            ])

        safety_factor = prediction.get('safety_factor')
        if safety_factor is not None:
            results_data.append([
                "Safety Factor",
                safety_factor,
                "N/A",
                "N/A",
                "-"
            ])

        # Write results
        for result_row in results_data:
            for col, value in enumerate(result_row, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = self.NORMAL_BORDER

                # Format numbers
                if col in [3, 4, 5] and isinstance(value, (int, float)):
                    if value != "N/A":
                        if isinstance(value, float):
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0'

            row += 1

        # Model coefficients section (if available)
        model_coeffs = prediction.get('model_coefficients')
        if model_coeffs:
            row += 1
            ws[f'B{row}'] = "Model Coefficients"
            ws[f'B{row}'].font = self.SUBTITLE_FONT
            row += 1

            headers = ["Coefficient", "Value", "Description"]
            for col, header in enumerate(headers, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.NORMAL_BORDER

            row += 1

            coeff_descriptions = {
                'A': 'Scaling factor',
                'alpha': 'Temperature swing exponent',
                'beta': 'Frequency exponent',
                'K': 'Scaling constant',
                'beta1': 'Delta Tj exponent',
                'beta2': 'Mean temperature factor',
                'beta3': 'On time exponent',
                'beta4': 'Current exponent',
                'beta5': 'Voltage exponent',
                'beta6': 'Diameter exponent',
                'Ea': 'Activation energy (eV)',
            }

            for coeff_name, coeff_value in model_coeffs.items():
                ws.cell(row=row, column=2, value=coeff_name)
                ws.cell(row=row, column=3, value=coeff_value)
                ws.cell(row=row, column=4, value=coeff_descriptions.get(coeff_name, ''))

                for col in range(2, 5):
                    ws.cell(row=row, column=col).border = self.NORMAL_BORDER

                row += 1

        # Adjust column widths
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10

        # Freeze header
        if self.config.freeze_headers:
            ws.freeze_panes = 'B5'

    def _create_cycles_sheet(self, wb: Workbook, cycles: List[Dict[str, Any]]):
        """Create the rainflow cycles sheet."""
        ws = wb.create_sheet("Rainflow Cycles")

        # Title
        ws['B1'] = "Rainflow Cycle Counting Results"
        ws['B1'].font = self.TITLE_FONT
        ws.merge_cells('B1:E1')

        row = 3

        # Summary statistics
        total_cycles = sum(c.get('cycles', 0) for c in cycles)
        ws[f'B{row}'] = "Total Cycle Counts:"
        ws[f'C{row}'] = total_cycles
        ws[f'C{row}'].number_format = '#,##0'
        row += 1

        ws[f'B{row}'] = "Unique Ranges:"
        ws[f'C{row}'] = len(cycles)
        row += 2

        # Table header
        headers = ["Stress Range", "Mean Value", "Cycles", "Damage Contribution"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.NORMAL_BORDER

        row += 1

        # Sort cycles by stress range descending
        sorted_cycles = sorted(cycles, key=lambda x: x.get('stress_range', 0), reverse=True)

        # Write cycle data
        for cycle in sorted_cycles:
            stress_range = cycle.get('stress_range', 0)
            mean_value = cycle.get('mean_value', 0)
            cycle_count = cycle.get('cycles', 0)
            damage = cycle.get('damage', 0)

            ws.cell(row=row, column=2, value=stress_range)
            ws.cell(row=row, column=3, value=mean_value)
            ws.cell(row=row, column=4, value=cycle_count)
            ws.cell(row=row, column=5, value=damage)

            for col in range(2, 6):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.NORMAL_BORDER

                if col in [2, 3]:  # Numeric columns
                    cell.number_format = '0.00'

            row += 1

        # Adjust column widths
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18

        # Freeze header
        if self.config.freeze_headers:
            ws.freeze_panes = 'B7'

    def _create_mission_profile_sheet(self, wb: Workbook, mission_profile: Dict[str, Any]):
        """Create the mission profile sheet."""
        ws = wb.create_sheet("Mission Profile")

        # Title
        ws['B1'] = "Mission Profile"
        ws['B1'].font = self.TITLE_FONT
        ws.merge_cells('B1:D1')

        row = 3

        # Mission profile data
        headers = ["Parameter", "Value", "Unit"]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.NORMAL_BORDER

        row += 1

        # Define mission profile parameter mapping
        profile_params = {
            'mission_name': ('Mission Name', ''),
            'mission_duration': ('Mission Duration', 'hours'),
            'cycles_per_mission': ('Cycles per Mission', ''),
            'operating_temperature': ('Operating Temperature', '°C'),
            'storage_temperature': ('Storage Temperature', '°C'),
            'altitude': ('Altitude', 'm'),
            'vibration_level': ('Vibration Level', 'g'),
            'humidity': ('Humidity', '%'),
        }

        for key, (name, unit) in profile_params.items():
            if key in mission_profile:
                value = mission_profile[key]
                formatted_value = self._format_value(value)

                ws.cell(row=row, column=2, value=name)
                ws.cell(row=row, column=3, value=formatted_value)
                ws.cell(row=row, column=4, value=unit)

                for col in range(2, 5):
                    ws.cell(row=row, column=col).border = self.NORMAL_BORDER

                row += 1

        # Add profile segments if available
        segments = mission_profile.get('segments', [])
        if segments:
            row += 1
            ws[f'B{row}'] = "Mission Segments"
            ws[f'B{row}'].font = self.SUBTITLE_FONT
            row += 1

            headers = ["Segment", "Duration", "Temperature", "Load"]
            for col, header in enumerate(headers, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = self.HEADER_ALIGNMENT
                cell.border = self.NORMAL_BORDER

            row += 1

            for i, segment in enumerate(segments, 1):
                ws.cell(row=row, column=2, value=i)
                ws.cell(row=row, column=3, value=segment.get('duration', 'N/A'))
                ws.cell(row=row, column=4, value=segment.get('temperature', 'N/A'))
                ws.cell(row=row, column=5, value=segment.get('load', 'N/A'))

                for col in range(2, 6):
                    ws.cell(row=row, column=col).border = self.NORMAL_BORDER

                row += 1

        # Adjust column widths
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    def _create_charts_sheet(
        self,
        wb: Workbook,
        prediction: Dict[str, Any],
        parameters: Dict[str, Any]
    ):
        """Create a sheet with embedded charts."""
        ws = wb.create_sheet("Charts")

        # Title
        ws['B1'] = "Analysis Charts"
        ws['B1'].font = self.TITLE_FONT

        # Create lifetime comparison chart (using data from summary)
        lifetime_years = prediction.get('predicted_lifetime_years')
        if lifetime_years is not None:
            # Add chart data
            row = 3
            ws[f'B{row}'] = "Category"
            ws[f'C{row}'] = "Years"
            row += 1
            ws[f'B{row}'] = "Predicted Lifetime"
            ws[f'C{row}'] = lifetime_years
            ws[f'C{row}'].number_format = '0.00'
            row += 1
            ws[f'B{row}'] = "5-Year Target"
            ws[f'C{row}'] = 5
            row += 1
            ws[f'B{row}'] = "10-Year Target"
            ws[f'C{row}'] = 10

            # Create bar chart
            chart = BarChart()
            chart.title = "Lifetime Comparison"
            chart.y_axis.title = "Years"
            chart.x_axis.title = "Category"
            chart.height = 10
            chart.width = 15

            data = Reference(ws, min_col=2, min_row=3, max_row=6, max_col=3)
            cats = Reference(ws, min_col=2, min_row=4, max_row=6)
            values = Reference(ws, min_col=3, min_row=4, max_row=6)

            chart.add_data(values, titles_from_data=False)
            chart.set_categories(cats)

            # Position chart
            ws.add_chart(chart, "B8")

        # Create damage accumulation chart if data available
        damage = prediction.get('total_damage')
        if damage is not None and lifetime_years is not None:
            # Add data for damage curve
            row = 15
            ws[f'E{row}'] = "Time (Years)"
            ws[f'F{row}'] = "Damage Ratio"
            row += 1

            # Generate damage curve data points
            for i in range(0, int(lifetime_years * 1.2) + 2, 2):
                damage_ratio = i / lifetime_years if lifetime_years > 0 else 0
                ws[f'E{row}'] = i
                ws[f'F{row}'] = min(1.2, damage_ratio)
                row += 1

            # Create line chart
            chart = LineChart()
            chart.title = "Damage Accumulation Over Time"
            chart.y_axis.title = "Damage Ratio"
            chart.x_axis.title = "Time (Years)"
            chart.height = 10
            chart.width = 15

            cats = Reference(ws, min_col=5, min_row=16, max_row=row - 1)
            values = Reference(ws, min_col=6, min_row=16, max_row=row - 1)

            chart.add_data(values, titles_from_data=False)
            chart.set_categories(cats)

            ws.add_chart(chart, "E8")

    def _format_model_name(self, model_type: str) -> str:
        """Format model type for display."""
        names = {
            "coffin-manson": "Coffin-Manson",
            "coffin-manson-arrhenius": "Coffin-Manson-Arrhenius",
            "norris-landzberg": "Norris-Landzberg",
            "cips-2008": "CIPS 2008 (Bayerer)",
            "lesit": "LESIT",
        }
        return names.get(model_type, model_type.replace('-', ' ').title())

    def _format_value(self, value: Any) -> str:
        """Format a value for display."""
        if isinstance(value, float):
            return f"{value:.4g}"
        elif isinstance(value, int):
            return f"{value:,}"
        else:
            return str(value)

    def _get_lifetime_status_text(self, value: float, units: str) -> str:
        """Get status text for lifetime value."""
        if units == 'years':
            if value < 1:
                return "CRITICAL"
            elif value < 5:
                return "WARNING"
            elif value < 10:
                return "NOTICE"
            else:
                return "OK"
        else:  # cycles
            if value < 10000:
                return "LOW"
            else:
                return "OK"

    def _get_damage_status_text(self, damage: float) -> str:
        """Get status text for damage value."""
        if damage >= 1.0:
            return "FAILURE"
        elif damage >= 0.8:
            return "HIGH"
        elif damage >= 0.5:
            return "MODERATE"
        else:
            return "LOW"


def generate_prediction_excel(
    prediction: Dict[str, Any],
    parameters: Dict[str, Any],
    mission_profile: Optional[Dict[str, Any]] = None,
    rainflow_cycles: Optional[List[Dict[str, Any]]] = None,
    output_path: Optional[str] = None,
    config: Optional[ExcelConfig] = None
) -> bytes:
    """
    Convenience function to generate an Excel report.

    Args:
        prediction: Prediction result data
        parameters: Model input parameters
        mission_profile: Mission profile data (optional)
        rainflow_cycles: Rainflow cycle counting results (optional)
        output_path: Optional file path to save the Excel file
        config: Optional Excel configuration

    Returns:
        Excel file content as bytes
    """
    generator = ExcelGenerator(config)
    return generator.generate_lifetime_report(
        prediction=prediction,
        parameters=parameters,
        mission_profile=mission_profile,
        rainflow_cycles=rainflow_cycles,
        output_path=output_path
    )
